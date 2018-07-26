# Excel

import openpyxl
import os

os.chdir('c:\\users\\drew\\downloads\\')

workbook = openpyxl.load_workbook('example.xlsx')
#commented out because they don't work.
# Need to figure out why they aren't working
#type(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')
#commented out because they don't work.
# Need to figure out why they aren't working
#type(sheet)

print(workbook.get_sheet_names())

cell = sheet['A1']
print(cell.value)
print(str(cell.value))
print(str(sheet['A1'].value))
print(sheet.cell(row=1, column=1))

cell1 = sheet['B1']
print(cell1.value)
print(str(cell1.value))
print(str(sheet['B1'].value))
print(sheet.cell(row=1, column=2))

print("Now we'll do a code that loops over multiple cells")
for i in range(1, 8):
    print(i, sheet.cell(row=1, column=2).value)
